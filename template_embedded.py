import base64
import io
import os
import tempfile
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side

def create_minimal_template():
    """Crea un template Excel minimo"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Template"
    
    # Impostazioni di base
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['H'].width = 15
    
    # Stili
    header_font = Font(bold=True, size=14)
    normal_font = Font(size=12)
    border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Etichetta 1
    ws['B4'] = "TRASPORTATORE"
    ws['B4'].font = header_font
    ws['B6'] = ""
    ws['B6'].font = normal_font
    ws['B6'].border = border
    
    ws['B12'] = "TARGA"
    ws['B12'].font = header_font
    ws['B14'] = ""
    ws['B14'].font = normal_font
    ws['B14'].border = border
    
    ws['H12'] = "ORA"
    ws['H12'].font = header_font
    ws['H14'] = ""
    ws['H14'].font = normal_font
    ws['H14'].border = border
    
    ws['B20'] = "VIAGGIO"
    ws['B20'].font = header_font
    ws['B22'] = ""
    ws['B22'].font = normal_font
    ws['B22'].border = border
    
    ws['H20'] = "DATA"
    ws['H20'].font = header_font
    ws['H22'] = ""
    ws['H22'].font = normal_font
    ws['H22'].border = border
    
    ws['B27'] = "SEQUENZA"
    ws['B27'].font = header_font
    ws['B29'] = ""
    ws['B29'].font = normal_font
    ws['B29'].border = border
    
    ws['H27'] = "TIPO"
    ws['H27'].font = header_font
    ws['H29'] = ""
    ws['H29'].font = normal_font
    ws['H29'].border = border
    
    # Etichetta 2
    ws['B36'] = "TRASPORTATORE"
    ws['B36'].font = header_font
    ws['B38'] = ""
    ws['B38'].font = normal_font
    ws['B38'].border = border
    
    ws['B44'] = "TARGA"
    ws['B44'].font = header_font
    ws['B46'] = ""
    ws['B46'].font = normal_font
    ws['B46'].border = border
    
    ws['H44'] = "ORA"
    ws['H44'].font = header_font
    ws['H46'] = ""
    ws['H46'].font = normal_font
    ws['H46'].border = border
    
    ws['B52'] = "VIAGGIO"
    ws['B52'].font = header_font
    ws['B54'] = ""
    ws['B54'].font = normal_font
    ws['B54'].border = border
    
    ws['H52'] = "DATA"
    ws['H52'].font = header_font
    ws['H54'] = ""
    ws['H54'].font = normal_font
    ws['H54'].border = border
    
    ws['B59'] = "SEQUENZA"
    ws['B59'].font = header_font
    ws['B61'] = ""
    ws['B61'].font = normal_font
    ws['B61'].border = border
    
    ws['H59'] = "TIPO"
    ws['H59'].font = header_font
    ws['H61'] = ""
    ws['H61'].font = normal_font
    ws['H61'].border = border
    
    return wb

def get_template_bytes():
    """Restituisce il template Excel come bytes"""
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        template_path = tmp.name
        wb = create_minimal_template()
        wb.save(template_path)
        with open(template_path, 'rb') as f:
            return f.read()

def get_template_file():
    """Restituisce il template Excel come file-like object"""
    return io.BytesIO(get_template_bytes())