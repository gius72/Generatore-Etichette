import streamlit as st
import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side
import re
import tempfile
import io
import base64

# Importa il template da GitHub
from s3_template import get_template_from_url

# ---------------------------------------------------------------------------
# Costanti ATL
# ---------------------------------------------------------------------------
ATL_ROW_HEIGHT = 35      # altezza righe 2-3 e 34-35 in pt
ATL_FONT_SIZE  = 52      # dimensione font "ATL"
ATL_BORDER = Border(
    left=Side(style="medium"),
    right=Side(style="medium"),
    top=Side(style="medium"),
    bottom=Side(style="none"),
)
ATL_BORDER_BOTTOM = Border(
    left=Side(style="medium"),
    right=Side(style="medium"),
    top=Side(style="none"),
    bottom=Side(style="medium"),
)

# ---------------------------------------------------------------------------
# Funzioni di filtro
# ---------------------------------------------------------------------------

def filtra_sap(df, area, rimorchio):
    """Filtra i dati SAP basandosi su area e rimorchio."""
    df = df.copy()
    df.drop_duplicates(subset="Viaggio", inplace=True)

    if "DescrSpedizioniere" in df.columns:
        df = df[df["DescrSpedizioniere"] != "NUMBER 1 LOGISTICS GROUP S.P.A."]

    if "Rimorchio" in df.columns:
        df["Rimorchio"] = df["Rimorchio"].fillna("Orario Fisso")
        df.loc[df["Rimorchio"].astype(str).str.strip() == "", "Rimorchio"] = "Orario Fisso"

    if area != "Tutti":
        if area == "Italia":
            df = df[df["Nazione Dest"] == "IT"]
        elif area == "Estero":
            df = df[df["Nazione Dest"] != "IT"]

    if rimorchio != "Tutti":
        if rimorchio == "A Piazzale":
            df = df[df["Rimorchio"].astype(str).str.contains("A Piazzale", na=False)]
        elif rimorchio == "Orario Fisso":
            df = df[df["Rimorchio"].astype(str).str.contains("Orario Fisso", na=False)]

    return df.reset_index(drop=True)


def filtra_dpe(df, tipo_ingaggio, tipo_gestione):
    """Filtra i dati DPE basandosi su tipo ingaggio e tipo gestione."""
    df = df.copy()

    veicolo_col        = next((c for c in df.columns if c.strip().lower() == "veicolo"), None)
    trasportatore_col  = next((c for c in df.columns if "trasportatore" in c.strip().lower()), None)
    tipo_ingaggio_col  = next((c for c in df.columns if "tipo" in c.strip().lower() and "ingaggio" in c.strip().lower()), None)
    tipo_gestione_col  = next((c for c in df.columns if "tipo" in c.strip().lower() and "gestione" in c.strip().lower()), None)
    dt_ingresso_prev_col = next((c for c in df.columns if "dt" in c.strip().lower() and "ingresso" in c.strip().lower() and "prev" in c.strip().lower()), None)
    targa_col          = next((c for c in df.columns if "targa" in c.strip().lower() and "rimorchio" in c.strip().lower()), None)
    viaggio_col        = next((c for c in df.columns if "viaggio" in c.strip().lower()), None)
    sequenza_col       = next((c for c in df.columns if "sequenza" in c.strip().lower()), None)

    if not tipo_ingaggio_col or not tipo_gestione_col or not dt_ingresso_prev_col:
        st.error("Colonne richieste non trovate nel file DPE.")
        return pd.DataFrame()

    if tipo_ingaggio != "Tutti":
        if tipo_ingaggio == "Viaggi":
            df = df[df[tipo_ingaggio_col].astype(str).str.contains("TRATTA", na=False, case=False)]
        elif tipo_ingaggio == "Spole":
            df = df[df[tipo_ingaggio_col].astype(str).str.contains("SPOLE", na=False, case=False)]
        elif tipo_ingaggio == "Rifugio":
            df = df[df[tipo_ingaggio_col].astype(str).str.contains("RIFUGIO", na=False, case=False)]

    if tipo_gestione != "Tutti":
        if tipo_gestione == "A Piazzale":
            df = df[df[tipo_gestione_col].astype(str).str.strip().str.upper() == "1 - A PIAZZALE"]
        elif tipo_gestione == "Orario Fisso":
            df = df[df[tipo_gestione_col].astype(str).str.strip().str.upper() == "2 - ORARIO FISSO"]

    rename_map = {}
    if trasportatore_col:  rename_map[trasportatore_col]    = "Trasportatore"
    if veicolo_col:        rename_map[veicolo_col]           = "Veicolo"
    if targa_col:          rename_map[targa_col]             = "Targa Rimorchio Eff."
    if viaggio_col:        rename_map[viaggio_col]           = "Viaggio"
    if sequenza_col:       rename_map[sequenza_col]          = "Sequenza"
    if tipo_gestione_col:  rename_map[tipo_gestione_col]     = "Tipo Gestione"
    if tipo_ingaggio_col:  rename_map[tipo_ingaggio_col]     = "Tipo Ingaggio"
    if dt_ingresso_prev_col: rename_map[dt_ingresso_prev_col] = "Dt. Ingresso Prev."

    df = df.rename(columns=rename_map)
    df = df.sort_values(by="Dt. Ingresso Prev.") if "Dt. Ingresso Prev." in df.columns else df
    return df.reset_index(drop=True)


# ---------------------------------------------------------------------------
# Funzione ATL
# ---------------------------------------------------------------------------

def carica_atl(atl_file):
    """
    Legge il file ATL (Excel o CSV) e restituisce un set di valori
    della colonna 'Nr. Viaggio' dove 'Carico Automatico' == 1.
    """
    try:
        if atl_file.name.lower().endswith(".csv"):
            try:
                atl_file.seek(0)
                df_atl = pd.read_csv(atl_file, sep=";", encoding="utf-8")
                if len(df_atl.columns) <= 1:
                    atl_file.seek(0)
                    df_atl = pd.read_csv(atl_file, sep=",", encoding="utf-8")
            except Exception:
                atl_file.seek(0)
                df_atl = pd.read_csv(atl_file, sep=";", encoding="cp1252")
        else:
            df_atl = pd.read_excel(atl_file)

        # Trova le colonne in modo flessibile (case-insensitive, strip spazi)
        nr_viaggio_col    = next((c for c in df_atl.columns if "nr" in c.strip().lower() and "viaggio" in c.strip().lower()), None)
        carico_auto_col   = next((c for c in df_atl.columns if "carico" in c.strip().lower() and "automatico" in c.strip().lower()), None)

        if nr_viaggio_col is None or carico_auto_col is None:
            st.error(f"File ATL: colonne non trovate. Colonne presenti: {list(df_atl.columns)}")
            return set()

        # Filtra solo le righe con Carico Automatico == 1
        df_atl_filtrato = df_atl[df_atl[carico_auto_col].astype(str).str.strip() == "1"]

        # Restituisce un set di Nr. Viaggio (come stringhe normalizzate)
        viaggi_atl = set(df_atl_filtrato[nr_viaggio_col].astype(str).str.strip())
        return viaggi_atl

    except Exception as e:
        st.error(f"Errore nel caricamento del file ATL: {e}")
        return set()


def applica_flag_atl(df, viaggi_atl):
    """
    Aggiunge la colonna 'is_atl' al dataframe confrontando
    la colonna 'Viaggio' con il set di viaggi ATL.
    """
    if "Viaggio" in df.columns and viaggi_atl:
        df["is_atl"] = df["Viaggio"].astype(str).str.strip().isin(viaggi_atl)
    else:
        df["is_atl"] = False
    return df


# ---------------------------------------------------------------------------
# Funzioni di supporto Excel
# ---------------------------------------------------------------------------

def elabora_numerazione(df):
    """Applica la numerazione pari/dispari per ottimizzare la stampa."""
    n = len(df)
    metà = n // 2
    dispari = list(range(1, metà * 2, 2))
    pari    = list(range(2, n * 2 + 1, 2))
    numerazione = dispari[:metà] + pari[:n - metà]
    df["Ordine"] = numerazione
    df.sort_values(by="Ordine", inplace=True)
    df.reset_index(drop=True, inplace=True)
    return df


def clean_excel_text(value):
    """Pulisce il testo per Excel rimuovendo caratteri indesiderati."""
    if pd.isna(value):
        return ""
    return str(value).replace('\r', '').replace('\n', '').strip()


def format_hhmm(value):
    """Formatta i valori di tempo in formato HH:MM."""
    if pd.isna(value) or value == "":
        return ""
    try:
        dt = pd.to_datetime(value, errors="coerce")
        if not pd.isna(dt):
            return dt.strftime("%H:%M")
        val = str(value).strip()
        if re.match(r"^\d{1,2}\.\d{2}(:\d{2})?$", val):
            val = val.replace(".", ":")
        if re.match(r"^\d{1,2}:\d{2}:\d{2}$", val):
            hh, mm, _ = val.split(":")
            return f"{hh.zfill(2)}:{mm}"
        if re.match(r"^\d{1,2}:\d{2}$", val):
            hh, mm = val.split(":")
            return f"{hh.zfill(2)}:{mm}"
        return val
    except Exception:
        return str(value).strip()


def format_ddmm(value):
    """Formatta le date in formato DD/MM."""
    if pd.isna(value) or value == "":
        return ""
    try:
        return pd.to_datetime(value, errors="coerce").strftime("%d/%m")
    except Exception:
        return str(value)


def set_spola_style(ws, cell):
    """Formatta la cella con lo stile SPOLA (grigio con testo bianco)."""
    ws[cell].value = "SPOLA"
    ws[cell].font      = Font(color="FFFFFF", bold=True, size=28)
    ws[cell].fill      = PatternFill("solid", fgColor="808080")
    ws[cell].alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)


def scrivi_atl_cell(ws, cella_top, cella_bottom, merge_range):
    """
    Scrive la cella ATL in F2:F3 (etichetta 1) o F34:F35 (etichetta 2).
    - Imposta altezza righe, unisce celle, scrive testo con font/bordi.
    """
    # Altezza righe
    top_row    = int(cella_top[1:])
    bottom_row = int(cella_bottom[1:])
    ws.row_dimensions[top_row].height    = ATL_ROW_HEIGHT
    ws.row_dimensions[bottom_row].height = ATL_ROW_HEIGHT

    # Merge e contenuto
    ws.merge_cells(merge_range)
    ws[cella_top].value     = "ATL"
    ws[cella_top].font      = Font(name="Arial", size=ATL_FONT_SIZE, bold=True)
    ws[cella_top].alignment = Alignment(horizontal="center", vertical="center")
    ws[cella_top].border    = ATL_BORDER
    ws[cella_bottom].border = ATL_BORDER_BOTTOM


# ---------------------------------------------------------------------------
# Generazione etichette
# ---------------------------------------------------------------------------

def create_labels_from_template(df, template_path, output_path, filtro_dpe_tipo_ingaggio):
    """Crea le etichette Excel dal template."""
    wb = load_workbook(template_path)
    ws_template = wb.active
    total = len(df)

    for i in range(0, total, 2):
        ws_new = wb.copy_worksheet(ws_template)
        ws_new.title = f"Etichette_{(i // 2) + 1}"

        # Flag ATL per le due etichette del foglio
        is_atl_1 = bool(df.iloc[i].get("is_atl", False))   if i     < total else False
        is_atl_2 = bool(df.iloc[i+1].get("is_atl", False)) if i + 1 < total else False

        # Stampa cella ATL se necessario
        if is_atl_1:
            scrivi_atl_cell(ws_new, "F2",  "F3",  "F2:F3")
        if is_atl_2:
            scrivi_atl_cell(ws_new, "F34", "F35", "F34:F35")

        # --- ETICHETTA 1 ---
        if i < total:
            row1 = df.iloc[i]
            if row1.get("_source", "SAP") == "SAP":  # SAP
                ws_new["B6"].value  = clean_excel_text(row1.get("DescrSpedizioniere", ""))
                ws_new["B14"].value = clean_excel_text(row1.get("Numero Targa", ""))
                ws_new["H14"].value = format_hhmm(row1.get("Ora Carico da", ""))
                ws_new["H14"].alignment = Alignment(wrap_text=False)
                ws_new["B22"].value = clean_excel_text(row1.get("Viaggio", ""))
                ws_new["H22"].value = format_ddmm(row1.get("Data Carico", ""))
                ws_new["B29"].value = f"{clean_excel_text(row1.get('Sequenza fermate', ''))} [{clean_excel_text(row1.get('Nazione Dest', ''))}]"
                ws_new["H29"].value = clean_excel_text(row1.get("Rimorchio", ""))
            else:  # DPE
                trasportatore = clean_excel_text(row1.get("Trasportatore", ""))
                veicolo       = clean_excel_text(row1.get("Veicolo", ""))
                if veicolo and veicolo.lower() != "nan":
                    trasportatore = f"{trasportatore} ({veicolo})"
                ws_new["B6"].value = trasportatore

                targa = clean_excel_text(row1.get("Targa Rimorchio Eff.", ""))
                if targa.lower() == "nan":
                    targa = ""
                ws_new["B14"].value = targa

                dt = row1.get("Dt. Ingresso Prev.", "")
                if "SPOLE" in str(row1.get("Tipo Ingaggio", "")).upper():
                    set_spola_style(ws_new, "H14")
                else:
                    ws_new["H14"].value = format_hhmm(dt)
                    ws_new["H14"].alignment = Alignment(wrap_text=False)

                ws_new["B22"].value = clean_excel_text(row1.get("Viaggio", ""))
                ws_new["H22"].value = format_ddmm(dt)
                ws_new["B29"].value = f"{clean_excel_text(row1.get('Sequenza', ''))} [IT]"

                tipo_gestione = clean_excel_text(row1.get("Tipo Gestione", ""))
                ws_new["H29"].value = "Orario Fisso" if tipo_gestione.strip().upper() == "2 - ORARIO FISSO" else "A Piazzale"

        # --- ETICHETTA 2 ---
        if i + 1 < total:
            row2 = df.iloc[i + 1]
            if row2.get("_source", "SAP") == "SAP":  # SAP
                ws_new["B38"].value = clean_excel_text(row2.get("DescrSpedizioniere", ""))
                ws_new["B46"].value = clean_excel_text(row2.get("Numero Targa", ""))
                ws_new["H46"].value = format_hhmm(row2.get("Ora Carico da", ""))
                ws_new["H46"].alignment = Alignment(wrap_text=False)
                ws_new["B54"].value = clean_excel_text(row2.get("Viaggio", ""))
                ws_new["H54"].value = format_ddmm(row2.get("Data Carico", ""))
                ws_new["H54"].alignment = Alignment(wrap_text=False)
                ws_new["B61"].value = f"{clean_excel_text(row2.get('Sequenza fermate', ''))} [{clean_excel_text(row2.get('Nazione Dest', ''))}]"
                ws_new["H61"].value = clean_excel_text(row2.get("Rimorchio", ""))
            else:  # DPE
                trasportatore = clean_excel_text(row2.get("Trasportatore", ""))
                veicolo       = clean_excel_text(row2.get("Veicolo", ""))
                if veicolo and veicolo.lower() != "nan":
                    trasportatore = f"{trasportatore} ({veicolo})"
                ws_new["B38"].value = trasportatore

                targa = clean_excel_text(row2.get("Targa Rimorchio Eff.", ""))
                if targa.lower() == "nan":
                    targa = ""
                ws_new["B46"].value = targa

                dt = row2.get("Dt. Ingresso Prev.", "")
                if "SPOLE" in str(row2.get("Tipo Ingaggio", "")).upper():
                    set_spola_style(ws_new, "H46")
                else:
                    ws_new["H46"].value = format_hhmm(dt)
                    ws_new["H46"].alignment = Alignment(wrap_text=False)

                ws_new["B54"].value = clean_excel_text(row2.get("Viaggio", ""))
                ws_new["H54"].value = format_ddmm(dt)
                ws_new["H54"].alignment = Alignment(wrap_text=False)
                ws_new["B61"].value = f"{clean_excel_text(row2.get('Sequenza', ''))} [IT]"

                tipo_gestione = clean_excel_text(row2.get("Tipo Gestione", ""))
                ws_new["H61"].value = "Orario Fisso" if tipo_gestione.strip().upper() == "2 - ORARIO FISSO" else "A Piazzale"

    try:
        wb.remove(ws_template)
        wb.save(output_path)
        return True, f"File etichette creato: {output_path}"
    except PermissionError:
        return False, f"Impossibile salvare '{output_path}'. Verifica che il file NON sia aperto e riprova."


def carica_file_csv_robusto(file):
    """Carica un file CSV provando diversi separatori e encoding."""
    separators = [";", ",", "\t", "|", " "]
    encodings  = ["utf-8", "cp1252", "iso-8859-1", "latin1"]

    for sep in separators:
        for enc in encodings:
            try:
                file.seek(0)
                df = pd.read_csv(file, encoding=enc, sep=sep)
                if len(df.columns) > 1:
                    return df
            except Exception:
                continue

    raise ValueError("Impossibile caricare il file CSV con i separatori e encoding disponibili")


# ---------------------------------------------------------------------------
# UI principale
# ---------------------------------------------------------------------------

def main():
    st.set_page_config(page_title="Generatore Etichette", layout="wide")
    st.title("Generatore Etichette")
    st.write("Carica i file SAP e DPE, scegli i filtri e genera le etichette in Excel.")

    col1, col2 = st.columns(2)

    with col1:
        st.subheader("File di input")
        sap_file      = st.file_uploader("Carica file SAP (Excel)",            type=["xlsx", "xls"])
        dpe_file      = st.file_uploader("Carica file DPE (Excel o CSV)",       type=["xlsx", "xls", "csv"])
        atl_file      = st.file_uploader("Carica file ATL (Excel o CSV) — opzionale", type=["xlsx", "xls", "csv"])
        template_file = st.file_uploader("Carica template etichette (opzionale)", type=["xlsx"])

    with col2:
        st.subheader("Filtri SAP")
        filtro_sap_area      = st.selectbox("Area SAP",       ["Tutti", "Italia", "Estero"])
        filtro_sap_rimorchio = st.selectbox("Rimorchio SAP",  ["Tutti", "A Piazzale", "Orario Fisso"])
        filtro_sap_atl       = st.selectbox("ATL SAP",        ["Tutti", "Solo ATL", "No ATL"],
                                             help="Filtra le etichette SAP in base alla presenza nel file ATL")

        st.subheader("Filtri DPE")
        filtro_dpe_tipo_ingaggio = st.selectbox("Tipo Ingaggio DPE", ["Tutti", "Viaggi", "Spole", "Rifugio"])
        filtro_dpe_tipo_gestione = st.selectbox("Tipo Gestione DPE", ["Tutti", "A Piazzale", "Orario Fisso"])
        filtro_dpe_atl           = st.selectbox("ATL DPE",           ["Tutti", "Solo ATL", "No ATL"],
                                                 help="Filtra le etichette DPE in base alla presenza nel file ATL")

    col3, col4 = st.columns(2)

    with col3:
        stampa_sap = st.checkbox("Stampa SAP", value=True)
        stampa_dpe = st.checkbox("Stampa DPE", value=True)

    with col4:
        output_path = st.text_input("Nome file di output", "etichette_generate.xlsx")

    if st.button("Genera Etichette", type="primary"):

        # Avviso se si usano filtri ATL senza file ATL
        usa_filtro_atl = (filtro_sap_atl != "Tutti" or filtro_dpe_atl != "Tutti")
        if usa_filtro_atl and atl_file is None:
            st.warning("Hai selezionato un filtro ATL ma non hai caricato il file ATL. Il filtro verrà ignorato.")

        # Caricamento template
        if template_file:
            with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
                template_path = tmp.name
                tmp.write(template_file.read())
        else:
            try:
                template_path = get_template_from_url()
            except Exception as e:
                st.error(f"Non è stato possibile scaricare il template: {e}")
                st.info("Carica manualmente un template usando il campo apposito.")
                return

        # Caricamento file ATL
        viaggi_atl = set()
        if atl_file is not None:
            viaggi_atl = carica_atl(atl_file)
            if viaggi_atl:
                st.info(f"File ATL caricato: {len(viaggi_atl)} viaggi con Carico Automatico = 1")
            else:
                st.warning("File ATL caricato ma nessun viaggio trovato con Carico Automatico = 1.")

        df_sap_filtered = pd.DataFrame()
        df_dpe_filtered = pd.DataFrame()

        # ----------------------------------------------------------------
        # Elaborazione SAP
        # ----------------------------------------------------------------
        if stampa_sap and sap_file is not None:
            try:
                df_sap = pd.read_excel(sap_file)
                df_sap_filtered = filtra_sap(df_sap, filtro_sap_area, filtro_sap_rimorchio)

                # Flag ATL
                df_sap_filtered = applica_flag_atl(df_sap_filtered, viaggi_atl)

                # Filtro ATL
                if filtro_sap_atl == "Solo ATL" and viaggi_atl:
                    df_sap_filtered = df_sap_filtered[df_sap_filtered["is_atl"]]
                elif filtro_sap_atl == "No ATL" and viaggi_atl:
                    df_sap_filtered = df_sap_filtered[~df_sap_filtered["is_atl"]]

                # Ordina per ora di carico
                if "Ora Carico da" in df_sap_filtered.columns:
                    df_sap_filtered = df_sap_filtered.sort_values(by="Ora Carico da")

                # Marca l'origine
                df_sap_filtered["_source"] = "SAP"

                st.info(f"Righe SAP dopo filtro: {len(df_sap_filtered)}")
            except Exception as e:
                st.error(f"Errore durante il caricamento del file SAP: {e}")
                return

        # ----------------------------------------------------------------
        # Elaborazione DPE
        # ----------------------------------------------------------------
        if stampa_dpe and dpe_file is not None:
            try:
                if dpe_file.name.endswith(".csv"):
                    df_dpe = carica_file_csv_robusto(dpe_file)
                else:
                    df_dpe = pd.read_excel(dpe_file)

                df_dpe_filtered = filtra_dpe(df_dpe, filtro_dpe_tipo_ingaggio, filtro_dpe_tipo_gestione)

                # Flag ATL
                df_dpe_filtered = applica_flag_atl(df_dpe_filtered, viaggi_atl)

                # Filtro ATL
                if filtro_dpe_atl == "Solo ATL" and viaggi_atl:
                    df_dpe_filtered = df_dpe_filtered[df_dpe_filtered["is_atl"]]
                elif filtro_dpe_atl == "No ATL" and viaggi_atl:
                    df_dpe_filtered = df_dpe_filtered[~df_dpe_filtered["is_atl"]]

                # Ordina per data ingresso
                if "Dt. Ingresso Prev." in df_dpe_filtered.columns:
                    df_dpe_filtered = df_dpe_filtered.sort_values(by="Dt. Ingresso Prev.")

                # Marca l'origine
                df_dpe_filtered["_source"] = "DPE"

                st.info(f"Righe DPE dopo filtro: {len(df_dpe_filtered)}")
            except Exception as e:
                st.error(f"Errore durante il caricamento del file DPE: {e}")
                return

        # ----------------------------------------------------------------
        # Concatenazione — ogni df è già ordinato, _source traccia l'origine
        # ----------------------------------------------------------------
        df_finale = pd.concat([df_sap_filtered, df_dpe_filtered], ignore_index=True)

        if df_finale.empty:
            st.error("Nessun dato da elaborare dopo i filtri.")
            return

        st.info(f"Totale righe dopo concatenazione: {len(df_finale)}")

        # Assicura che is_atl esista sempre
        if "is_atl" not in df_finale.columns:
            df_finale["is_atl"] = False

        # Numerazione pari/dispari
        st.info("Applico numerazione pari/dispari...")
        df_finale = elabora_numerazione(df_finale)

        # Genera etichette
        with st.spinner("Generazione etichette in corso..."):
            success, msg = create_labels_from_template(df_finale, template_path, output_path, filtro_dpe_tipo_ingaggio)

        if success:
            st.success(msg)
            with open(output_path, "rb") as file:
                st.download_button(
                    label="Scarica file etichette",
                    data=file,
                    file_name=output_path,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
        else:
            st.error(msg)


if __name__ == "__main__":
    main()
